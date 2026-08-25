#!/usr/bin/env python3
"""
MacroSuite — Nutrition Planning Software
Dark Apple-style UI · Excel database · Automatic backup · Persistent settings.
"""

import sys
import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple
from dataclasses import dataclass, field
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QLineEdit, QDialog, QFormLayout, QComboBox, QDoubleSpinBox,
    QMessageBox, QCompleter, QHeaderView, QListWidget, QFrame,
    QStatusBar, QFileDialog, QAbstractItemView, QGridLayout,
    QInputDialog,
)
from PySide6.QtCore import Qt, QTimer, Signal, QSettings, QSize, QEvent, QObject, QPointF
from PySide6.QtGui import (
    QFont as QFontGui, QColor, QAction, QPixmap, QPainter, QPolygonF, QIcon,
)


# ━━━━━━━━━━━━━━━━━━━━━━ CONSTANTS ━━━━━━━━━━━━━━━━━━━━━━

APP_TITLE = "MacroSuite — Nutrition Planning Software"

# Internal nutrition field keys and their default UI labels (used in dialogs)
NUTRITION_FIELDS = [
    ("energy_kj",    "Energy (kJ)"),
    ("energy_kcal",  "Energy (kcal)"),
    ("fat",          "Fat (g)"),
    ("saturated_fat","Sat. Fat (g)"),
    ("carbohydrate", "Carbs (g)"),
    ("sugars",       "Sugars (g)"),
    ("fibre",        "Fibre (g)"),
    ("protein",      "Protein (g)"),
    ("salt",         "Salt (g)"),
]
NUTRITION_KEYS   = [k for k, _ in NUTRITION_FIELDS]
NUTRITION_LABELS = [l for _, l in NUTRITION_FIELDS]

# Excel column mapping: which internal key lives at which column index (0-based in row tuple)
# This MUST match the actual database column layout exactly.
# Col A(0)=ID, B(1)=human verified, C(2)=Name, D(3)=Brand, E(4)=Product, F(5)=Basis,
# G(6)=Energy(kJ), H(7)=Energy(kcal), I(8)=Fat, J(9)=SatFat, K(10)=Carbs,
# L(11)=Sugars, M(12)=Fibre, N(13)=Protein, O(14)=Salt, P(15)=PackageSize
EXCEL_NUTRITION_START_COL = 7  # 1-based column index where nutrition starts (col G)
EXCEL_COL_ORDER = [
    "energy_kj",     # col 7  (G) — exactly as in database
    "energy_kcal",   # col 8  (H) — exactly as in database
    "fat",           # col 9  (I)
    "saturated_fat", # col 10 (J)
    "carbohydrate",  # col 11 (K)
    "sugars",        # col 12 (L)
    "fibre",         # col 13 (M)
    "protein",       # col 14 (N)
    "salt",          # col 15 (O)
]

SETTINGS_ORG     = "MacroSuite"
SETTINGS_APP     = "MacroSuite"
SETTINGS_LAST_DB = "last_database_path"

# ── Calorie text colors (light, comfortable on dark background) ──
CAL_TEXT_GREEN  = QColor(110, 210, 140)   # 0–149 kcal
CAL_TEXT_YELLOW = QColor(240, 220, 80)   # 150–399 kcal
CAL_TEXT_ORANGE = QColor(230, 140, 80)   # ≥400 kcal

# ── Theme colors ──
C_BG        = "#1c1c1e"
C_CARD      = "#2c2c2e"
C_CARD2     = "#323234"
C_BORDER    = "#38383a"
C_BORDER2   = "#48484a"
C_TEXT       = "#f5f5f7"
C_TEXT2      = "#98989d"
C_TEXT3      = "#636366"
C_ACCENT     = "#0a84ff"
C_ACCENT_HV  = "#409cff"
C_GREEN      = "#30d158"
C_RED        = "#ff453a"
C_TOTAL_BG   = "#1a3a5c"
C_PER100_BG  = "#2a2a2c"


# ━━━━━━━━━━━━━━━ AUTO-SELECT ON FOCUS ━━━━━━━━━━━━━━━━

class SelectAllOnFocus(QObject):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._just_focused = set()

    def eventFilter(self, obj, event):
        et = event.type()

        if et == QEvent.KeyPress:
            ctrl_space = event.key() == Qt.Key_Space and event.modifiers() & Qt.ControlModifier
            alt_down   = event.key() == Qt.Key_Down  and event.modifiers() & Qt.AltModifier
            if ctrl_space or alt_down:
                combo = None
                if isinstance(obj, QComboBox):
                    combo = obj
                elif isinstance(obj, QLineEdit) and isinstance(obj.parent(), QComboBox):
                    combo = obj.parent()
                if combo:
                    combo.showPopup()
                    return True

            if event.key() == Qt.Key_Tab and isinstance(obj, QLineEdit):
                comp = obj.completer()
                if comp and comp.popup() and comp.popup().isVisible():
                    popup = comp.popup()
                    model = comp.completionModel()
                    if model.rowCount() > 0:
                        cur = popup.currentIndex()
                        nxt = (cur.row() + 1) if cur.isValid() and cur.row() >= 0 else 0
                        if nxt >= model.rowCount():
                            nxt = 0
                        idx = model.index(nxt, 0)
                        popup.setCurrentIndex(idx)
                        text = idx.data()
                        if text:
                            obj.setText(text)
                    return True

        target = obj
        if isinstance(obj, QComboBox) and obj.isEditable() and obj.lineEdit():
            target = obj.lineEdit()

        if not isinstance(target, (QDoubleSpinBox, QLineEdit)):
            return super().eventFilter(obj, event)
        if isinstance(target, QLineEdit) and target.isReadOnly():
            return super().eventFilter(obj, event)

        oid = id(target)

        if et == QEvent.FocusIn:
            self._just_focused.add(oid)
            QTimer.singleShot(0, lambda o=target: self._sel(o))
        elif et == QEvent.MouseButtonRelease:
            if oid in self._just_focused:
                self._just_focused.discard(oid)
                QTimer.singleShot(0, lambda o=target: self._sel(o))
        elif et == QEvent.FocusOut:
            self._just_focused.discard(oid)

        return super().eventFilter(obj, event)

    @staticmethod
    def _sel(obj):
        if isinstance(obj, QDoubleSpinBox):
            obj.selectAll()
        elif isinstance(obj, QLineEdit) and not obj.isReadOnly():
            obj.selectAll()


# ━━━━━━━━━━━━━━━━ ARROW ICON HELPER ━━━━━━━━━━━━━━━━━

def _create_arrow_image() -> str:
    path = os.path.join(tempfile.gettempdir(), "macrosuite_arrow.png")
    pm = QPixmap(14, 10)
    pm.fill(Qt.transparent)
    p = QPainter(pm)
    p.setRenderHint(QPainter.Antialiasing)
    p.setPen(Qt.NoPen)
    p.setBrush(QColor(C_TEXT2))
    p.drawPolygon(QPolygonF([QPointF(1, 1), QPointF(13, 1), QPointF(7, 9)]))
    p.end()
    pm.save(path, "PNG")
    return path


# ━━━━━━━━━━━━━━━━━━ APPLE DARK THEME ━━━━━━━━━━━━━━━━━━

DARK_STYLE_TEMPLATE = """
* {{
    font-family: -apple-system, "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
}}
QMainWindow {{ background-color: {bg}; }}
QWidget     {{ background-color: {bg}; color: {tx}; }}
QFrame#dbBar {{
    background-color: {cd}; border-bottom: 1px solid {bd}; padding: 8px 16px;
}}
QFrame#dbBar QLabel {{ color: {tx2}; font-size: 12px; background: transparent; }}
QFrame#dbBar QLabel#dbPath {{ color: {tx}; font-size: 13px; font-weight: 500; background: transparent; }}
QTabWidget::pane {{ border: none; background: {bg}; }}
QTabBar {{ background: {bg}; }}
QTabBar::tab {{
    background: transparent; color: {tx2};
    padding: 12px 32px; border: none;
    border-bottom: 2px solid transparent;
    font-size: 13px; font-weight: 600;
}}
QTabBar::tab:selected {{ color: {ac}; border-bottom: 2px solid {ac}; }}
QTabBar::tab:hover:!selected {{ color: #c7c7cc; }}
QTableWidget {{
    gridline-color: {bd}; background-color: {cd};
    alternate-background-color: {cd2};
    selection-background-color: {ac}33; selection-color: {tx};
    border: 1px solid {bd}; border-radius: 10px;
    font-size: 12px; outline: none;
}}
QTableWidget::item {{ padding: 5px 10px; border: none; }}
QTableWidget::item:selected {{ background-color: {ac}44; }}
QHeaderView::section {{
    background-color: #3a3a3c; color: {tx2};
    padding: 8px 10px; border: none;
    border-right: 1px solid {bd2}; border-bottom: 1px solid {bd2};
    font-weight: 600; font-size: 11px;
}}
QHeaderView::section:first {{ border-top-left-radius: 10px; }}
QHeaderView::section:last  {{ border-top-right-radius: 10px; border-right: none; }}
QPushButton {{
    background-color: {ac}; color: #ffffff; border: none;
    padding: 8px 22px; border-radius: 8px;
    font-weight: 600; font-size: 13px; min-height: 18px;
}}
QPushButton:hover    {{ background-color: {achv}; }}
QPushButton:pressed  {{ background-color: #0071e3; }}
QPushButton:disabled {{ background-color: {bd2}; color: {tx3}; }}
QPushButton[class="danger"]       {{ background-color: {rd}; }}
QPushButton[class="danger"]:hover {{ background-color: #ff6961; }}
QPushButton[class="secondary"]       {{ background-color: #3a3a3c; color: {tx}; border: 1px solid {bd2}; }}
QPushButton[class="secondary"]:hover {{ background-color: {bd2}; }}
QPushButton[class="dbButton"]       {{ background: #3a3a3c; color: {tx}; border: 1px solid {bd2}; padding: 6px 16px; font-size: 12px; border-radius: 6px; }}
QPushButton[class="dbButton"]:hover {{ background: {bd2}; }}
QLineEdit {{
    padding: 8px 14px; border: 1px solid {bd2};
    border-radius: 8px; background: #3a3a3c; color: {tx}; font-size: 13px;
    selection-background-color: {ac};
}}
QLineEdit:focus {{ border-color: {ac}; }}
QLineEdit::placeholder {{ color: {tx3}; }}
QComboBox {{
    padding: 8px 36px 8px 14px;
    border: 1px solid {bd2}; border-radius: 8px;
    background: #3a3a3c; color: {tx}; font-size: 13px;
    min-width: 180px;
}}
QComboBox:focus {{ border-color: {ac}; }}
QComboBox::drop-down {{
    subcontrol-origin: padding; subcontrol-position: center right;
    width: 32px; border: none; border-left: 1px solid {bd2};
}}
QComboBox::down-arrow {{
    image: url({arrow});
    width: 14px; height: 10px;
}}
QComboBox QAbstractItemView {{
    background: #3a3a3c; color: {tx};
    border: 1px solid {bd2}; border-radius: 8px;
    selection-background-color: {ac}; outline: none;
}}
QDoubleSpinBox, QSpinBox {{
    padding: 8px 14px; border: 1px solid {bd2};
    border-radius: 8px; background: #3a3a3c; color: {tx}; font-size: 13px;
}}
QDoubleSpinBox:focus, QSpinBox:focus {{ border-color: {ac}; }}
QListWidget {{
    border: 1px solid {bd}; border-radius: 10px;
    background: {cd}; font-size: 13px; outline: none;
}}
QListWidget::item {{
    padding: 12px 16px; border-bottom: 1px solid {bd}; border-radius: 0;
}}
QListWidget::item:selected {{ background-color: {ac}33; color: {ac}; }}
QListWidget::item:hover:!selected {{ background-color: #3a3a3c; }}
QListWidget::item:last {{ border-bottom: none; }}
QLabel {{ background: transparent; }}
QLabel#sectionTitle {{ font-size: 20px; font-weight: 700; color: {tx}; }}
QStatusBar {{
    background: {cd}; border-top: 1px solid {bd}; color: {tx3}; font-size: 12px;
}}
QFrame#separator {{ background-color: {bd}; max-height: 1px; }}
QScrollBar:vertical {{ background: {cd}; width: 8px; border-radius: 4px; }}
QScrollBar::handle:vertical {{ background: {tx3}; border-radius: 4px; min-height: 30px; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{ background: {cd}; height: 8px; border-radius: 4px; }}
QScrollBar::handle:horizontal {{ background: {tx3}; border-radius: 4px; min-width: 30px; }}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}
QMessageBox {{ background-color: {cd}; }}
QMessageBox QLabel {{ color: {tx}; }}
QDialog {{ background-color: {cd}; }}
QInputDialog {{ background-color: {cd}; }}
"""


def _build_stylesheet(arrow_path: str) -> str:
    return DARK_STYLE_TEMPLATE.format(
        bg=C_BG, cd=C_CARD, cd2=C_CARD2, bd=C_BORDER, bd2=C_BORDER2,
        tx=C_TEXT, tx2=C_TEXT2, tx3=C_TEXT3,
        ac=C_ACCENT, achv=C_ACCENT_HV, rd=C_RED,
        arrow=arrow_path.replace("\\", "/"),
    )


# ━━━━━━━━━━━━━━━━━━━━ DATA CLASSES ━━━━━━━━━━━━━━━━━━━━

@dataclass
class Ingredient:
    id: int
    name: str
    brand: str = ""
    product_name: str = ""
    energy_kj: float = 0.0
    energy_kcal: float = 0.0
    fat: float = 0.0
    saturated_fat: float = 0.0
    carbohydrate: float = 0.0
    sugars: float = 0.0
    fibre: float = 0.0
    protein: float = 0.0
    salt: float = 0.0
    package_size: str = ""

    def scaled_nutrition(self, grams: float) -> Dict[str, float]:
        f = grams / 100.0
        return {k: round(getattr(self, k) * f, 2) for k in NUTRITION_KEYS}


@dataclass
class MealIngredient:
    ingredient_id: int
    amount_grams: float


@dataclass
class Meal:
    name: str
    items: List[MealIngredient] = field(default_factory=list)


@dataclass
class MenuEntry:
    item_type: str
    item_name: str
    amount: float


@dataclass
class Menu:
    name: str
    items: List[MenuEntry] = field(default_factory=list)


# ━━━━━━━━━━━━━━━━━━ DATABASE MANAGER ━━━━━━━━━━━━━━━━━━

class DatabaseManager:
    INGREDIENTS_SHEET = "Ingrediants"
    MEALS_SHEET       = "Meals"
    MENUS_SHEET       = "Menues"

    # Header keywords → internal field names (case-insensitive matching)
    _NUTRITION_DETECT = {
        "kcal":    "energy_kcal",
        "kj":      "energy_kj",
    }
    # Order matters: "saturated" must be checked before generic "fat"
    _NUTRITION_DETECT_ORDERED = [
        ("saturated", "saturated_fat"),
        ("carbohydrate", "carbohydrate"),
        ("sugar",    "sugars"),
        ("fibre",    "fibre"),
        ("fiber",    "fibre"),
        ("protein",  "protein"),
        ("salt",     "salt"),
        ("fat",      "fat"),  # must be AFTER "saturated"
    ]

    def __init__(self, path: str):
        self.path = Path(path)
        backup = self.path.parent / f"{self.path.stem}_backup{self.path.suffix}"
        if not backup.exists():
            shutil.copy2(self.path, backup)
        # Column index mapping — detected from headers on first load
        self._col_map: Dict[str, int] = {}  # field_name → 0-based column index
        self._max_col: int = 16

    def _detect_columns(self, header_row):
        """Detect which column index holds which nutrition field, from header text."""
        self._col_map = {}
        self._max_col = len(header_row)
        for idx, h in enumerate(header_row):
            if h is None:
                continue
            hl = str(h).lower()
            # Check kcal/kJ first (unique keywords)
            for keyword, field in self._NUTRITION_DETECT.items():
                if keyword in hl and field not in self._col_map:
                    self._col_map[field] = idx
                    break
            else:
                # Check ordered keywords
                for keyword, field in self._NUTRITION_DETECT_ORDERED:
                    if keyword in hl and field not in self._col_map:
                        self._col_map[field] = idx
                        break

        # Fallback: if we found fewer than 5 fields, use standard 16-col positions
        if len(self._col_map) < 5:
            self._col_map = {
                "energy_kj": 6, "energy_kcal": 7, "fat": 8,
                "saturated_fat": 9, "carbohydrate": 10, "sugars": 11,
                "fibre": 12, "protein": 13, "salt": 14,
            }

    def _get_col(self, field: str) -> int:
        return self._col_map.get(field, -1)

    @property
    def nutrition_display_order(self) -> List[str]:
        """Nutrition keys sorted by database column position."""
        ordered = sorted(
            [(idx, k) for k, idx in self._col_map.items() if k in set(NUTRITION_KEYS)],
            key=lambda x: x[0]
        )
        return [k for _, k in ordered] if ordered else NUTRITION_KEYS

    @property
    def nutrition_display_labels(self) -> List[str]:
        """Nutrition labels in database column order."""
        label_map = dict(NUTRITION_FIELDS)
        return [label_map.get(k, k) for k in self.nutrition_display_order]

    # ── Read headers exactly from database ──

    def load_headers(self) -> Dict[str, List[str]]:
        """Read row 1 headers from every sheet — exact text, no changes."""
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        result = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            result[sn] = [str(v) if v is not None else "" for v in row1]
        wb.close()
        return result

    # ── Read data ──

    def load_ingredients(self) -> Dict[int, Ingredient]:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        ws = wb[self.INGREDIENTS_SHEET]

        # Read header row and detect column positions
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        self._detect_columns(header_row)
        self._max_col = len(header_row)

        result = {}
        for row in ws.iter_rows(min_row=2, max_col=self._max_col, values_only=True):
            rid, name = row[0], row[2] if len(row) > 2 else None
            if rid is None or not name:
                continue

            def _val(field):
                """Safely get a value from the row at the detected column."""
                idx = self._col_map.get(field, -1)
                if 0 <= idx < len(row):
                    return row[idx]
                return None

            result[int(rid)] = Ingredient(
                id=int(rid), name=str(name).strip(),
                brand=str(row[3] or "").strip() if len(row) > 3 else "",
                product_name=str(row[4] or "").strip() if len(row) > 4 else "",
                energy_kj=_f(_val("energy_kj")),
                energy_kcal=_f(_val("energy_kcal")),
                fat=_f(_val("fat")),
                saturated_fat=_f(_val("saturated_fat")),
                carbohydrate=_f(_val("carbohydrate")),
                sugars=_f(_val("sugars")),
                fibre=_f(_val("fibre")),
                protein=_f(_val("protein")),
                salt=_f(_val("salt")),
                package_size=str(row[self._max_col - 1] or "").strip() if self._max_col > 15 and len(row) > self._max_col - 1 else "",
            )
        wb.close()
        return result

    def load_meals(self) -> Dict[str, Meal]:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        ws = wb[self.MEALS_SHEET]
        meals: Dict[str, Meal] = OrderedDict()

        # Read header to detect columns
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
        max_col = len(header)

        # New structure: C(2)=MealName, D(3)=IngredientName, G(6)=Basis(amount)
        for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
            meal_name = row[2] if len(row) > 2 else None
            if not meal_name:
                continue
            meal_name = str(meal_name).strip()
            if meal_name not in meals:
                meals[meal_name] = Meal(name=meal_name)

            ing_name = row[3] if len(row) > 3 else None
            if not ing_name:
                continue
            ing_name = str(ing_name).strip()

            # Parse amount from Basis column (col G) e.g. "200 g", "per 100 g"
            basis = row[6] if len(row) > 6 else None
            amount = _parse_basis(basis)

            # Find ingredient ID by name
            # (stored temporarily as -1; resolved against loaded ingredients later)
            ing_id = self._find_ingredient_id_by_name(ing_name)
            if ing_id >= 0:
                meals[meal_name].items.append(
                    MealIngredient(ingredient_id=ing_id, amount_grams=amount)
                )

        wb.close()
        return meals

    def _find_ingredient_id_by_name(self, name: str) -> int:
        """Look up ingredient ID by name from the ingredients sheet."""
        if not hasattr(self, '_ing_name_cache'):
            self._ing_name_cache = {}
            wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
            ws = wb[self.INGREDIENTS_SHEET]
            for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
                if row[0] is not None and row[2]:
                    self._ing_name_cache[str(row[2]).strip().lower()] = int(row[0])
            wb.close()
        return self._ing_name_cache.get(name.lower(), -1)

    def load_menus(self, ingredients: Dict[int, Ingredient] = None,
                   meals: Dict[str, Meal] = None) -> Dict[str, Menu]:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        ws = wb[self.MENUS_SHEET]
        menus: Dict[str, Menu] = OrderedDict()

        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
        max_col = len(header)

        # Collect ingredient and meal names for auto-detection
        ing_names = set()
        if ingredients:
            ing_names = {i.name.lower() for i in ingredients.values()}
        meal_names = set()
        if meals:
            meal_names = {n.lower() for n in meals.keys()}

        # New structure: C(2)=MenuName, D(3)=Ingredient/MealName, G(6)=Basis
        for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
            menu_name = row[2] if len(row) > 2 else None
            if not menu_name:
                continue
            menu_name = str(menu_name).strip()
            if menu_name not in menus:
                menus[menu_name] = Menu(name=menu_name)

            item_name = row[3] if len(row) > 3 else None
            if not item_name:
                continue
            item_name = str(item_name).strip()

            basis = row[6] if len(row) > 6 else None
            amount = _parse_basis(basis)

            # Auto-detect type: check meals first, then ingredients
            if item_name.lower() in meal_names:
                item_type = "meal"
            elif item_name.lower() in ing_names:
                item_type = "ingredient"
            else:
                item_type = "ingredient"  # default

            menus[menu_name].items.append(
                MenuEntry(item_type=item_type, item_name=item_name, amount=amount)
            )

        wb.close()
        return menus

    # ── Write (ABSOLUTE ZERO formatting changes, ZERO header changes) ──

    @staticmethod
    def _safe_write(ws, row, col, value):
        if value is not None and value != "":
            ws.cell(row, col).value = value
        elif (row, col) in ws._cells:
            ws._cells[(row, col)].value = None

    @staticmethod
    def _clear_existing_data(ws, min_row=2):
        for (r, c), cell in list(ws._cells.items()):
            if r >= min_row and cell.value is not None:
                cell.value = None

    @staticmethod
    def _copy_style_to(ws, row, col, style):
        if style is not None and (row, col) in ws._cells:
            ws._cells[(row, col)]._style = style

    def save_all(self, ingredients, meals, menus):
        wb = openpyxl.load_workbook(self.path)

        # ── INGREDIENTS: values only, NEVER touch headers or col B ──
        ws = wb[self.INGREDIENTS_SHEET]

        id_to_row: Dict[int, int] = {}
        for (r, c), cell in ws._cells.items():
            if r >= 2 and c == 1 and cell.value is not None:
                try:
                    id_to_row[int(cell.value)] = r
                except (ValueError, TypeError):
                    pass

        used_rows = set(id_to_row.values())
        written_ids = set()

        for ing in ingredients.values():
            if ing.id in id_to_row:
                r = id_to_row[ing.id]
            else:
                r = 2
                while r in used_rows:
                    r += 1
                used_rows.add(r)
                self._safe_write(ws, r, 1, ing.id)

            written_ids.add(ing.id)
            self._safe_write(ws, r, 3, ing.name)
            self._safe_write(ws, r, 4, ing.brand or None)
            self._safe_write(ws, r, 5, ing.product_name or None)
            self._safe_write(ws, r, 6, "per 100 g")
            # Write nutrition to DETECTED column positions (1-based)
            for field in NUTRITION_KEYS:
                col_idx = self._get_col(field)
                if col_idx >= 0:
                    val = getattr(ing, field)
                    self._safe_write(ws, r, col_idx + 1, val if val else None)
            self._safe_write(ws, r, self._max_col, ing.package_size or None)

        for old_id, old_row in id_to_row.items():
            if old_id not in written_ids:
                for col in range(3, 17):
                    if (old_row, col) in ws._cells:
                        ws._cells[(old_row, col)].value = None

        # ── MEALS: new structure — C=MealName, D=IngredientName, E=Brand, F=Product, G=Basis, H+=nutrition ──
        ws2 = wb[self.MEALS_SHEET]
        data_style = None
        for (r, c), cell in ws2._cells.items():
            if r >= 2:
                data_style = cell._style
                break
        self._clear_existing_data(ws2, min_row=2)
        r = 2
        for meal in meals.values():
            if not meal.items:
                # Empty meal — just write name in col C
                ws2.cell(r, 3).value = meal.name
                self._copy_style_to(ws2, r, 3, data_style)
                r += 1; continue
            for mi in meal.items:
                ing = ingredients.get(mi.ingredient_id)
                ws2.cell(r, 3).value = meal.name                          # C: Meal Name
                ws2.cell(r, 4).value = ing.name if ing else "?"           # D: Ingredient Name
                ws2.cell(r, 5).value = (ing.brand if ing else "") or None # E: Brand
                ws2.cell(r, 6).value = (ing.product_name if ing else "") or None  # F: Product
                ws2.cell(r, 7).value = f"{mi.amount_grams:.0f} g"        # G: Basis (amount)
                if ing:
                    nutr = ing.scaled_nutrition(mi.amount_grams)
                    # H+ nutrition — detect column order from Meals sheet headers
                    # Use positions 8-16 (1-based) matching the sheet layout
                    meal_nutr_cols = [8, 9, 10, 11, 12, 13, 14, 15, 16]
                    for c, key in zip(meal_nutr_cols, NUTRITION_KEYS):
                        ws2.cell(r, c).value = nutr.get(key, 0)
                max_c = max(ws2.max_column or 17, 17)
                for c in range(1, max_c + 1):
                    self._copy_style_to(ws2, r, c, data_style)
                r += 1

        # ── MENUS: new structure — C=MenuName, D=Ingredient/MealName, G=Basis, H+=nutrition ──
        ws3 = wb[self.MENUS_SHEET]
        data_style3 = None
        for (r, c), cell in ws3._cells.items():
            if r >= 2:
                data_style3 = cell._style
                break
        self._clear_existing_data(ws3, min_row=2)
        r = 2
        for menu in menus.values():
            if not menu.items:
                ws3.cell(r, 3).value = menu.name
                self._copy_style_to(ws3, r, 3, data_style3)
                r += 1; continue
            for entry in menu.items:
                ws3.cell(r, 3).value = menu.name        # C: Menu Name
                ws3.cell(r, 4).value = entry.item_name  # D: Ingredient/Meal Name

                # Calculate nutrition for this entry
                nutr = {}
                weight = 0
                if entry.item_type == "ingredient":
                    ing = _find_ing(entry.item_name, ingredients)
                    if ing:
                        ws3.cell(r, 5).value = ing.brand or None
                        ws3.cell(r, 6).value = ing.product_name or None
                        ws3.cell(r, 7).value = f"{entry.amount:.0f} g"
                        nutr = ing.scaled_nutrition(entry.amount)
                elif entry.item_type == "meal":
                    meal = meals.get(entry.item_name)
                    if meal:
                        ws3.cell(r, 7).value = f"{entry.amount:.0f} g"
                        mg, mt = NutritionCalc.meal_totals(meal, ingredients)
                        ratio = entry.amount / mg if mg > 0 else 0
                        nutr = {k: mt[k] * ratio for k in NUTRITION_KEYS}

                # Write nutrition H+ (cols 8-16, 1-based)
                menu_nutr_cols = [8, 9, 10, 11, 12, 13, 14, 15, 16]
                for c, key in zip(menu_nutr_cols, NUTRITION_KEYS):
                    ws3.cell(r, c).value = round(nutr.get(key, 0), 2)

                max_c = max(ws3.max_column or 17, 17)
                for c in range(1, max_c + 1):
                    self._copy_style_to(ws3, r, c, data_style3)
                r += 1

        tmp = self.path.with_suffix(".tmp.xlsx")
        try:
            wb.save(tmp); wb.close()
            if self.path.exists():
                self.path.unlink()
            tmp.rename(self.path)
        except Exception:
            wb.close()
            if tmp.exists():
                tmp.unlink()
            raise


def _f(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        try:
            return float(str(val).replace(",", ".").strip())
        except (ValueError, TypeError):
            return 0.0


def _parse_basis(basis) -> float:
    """Extract grams from basis string like 'per 100 g', '200 g', '50'."""
    if basis is None:
        return 100.0
    import re
    nums = re.findall(r'[\d.]+', str(basis))
    if nums:
        return float(nums[-1])
    return 100.0


# ━━━━━━━━━━━━━━━━━━ NUTRITION CALC ━━━━━━━━━━━━━━━━━━━━

class NutritionCalc:
    @staticmethod
    def meal_totals(meal, ingredients):
        total_g = 0.0
        totals = {k: 0.0 for k in NUTRITION_KEYS}
        for mi in meal.items:
            ing = ingredients.get(mi.ingredient_id)
            if not ing: continue
            total_g += mi.amount_grams
            for k, v in ing.scaled_nutrition(mi.amount_grams).items():
                totals[k] += v
        return round(total_g, 1), {k: round(v, 2) for k, v in totals.items()}

    @staticmethod
    def per100(total_g, totals):
        if total_g <= 0:
            return {k: 0.0 for k in NUTRITION_KEYS}
        return {k: round(v / total_g * 100, 2) for k, v in totals.items()}

    @staticmethod
    def menu_totals(menu, meals, ingredients):
        total_g = 0.0
        totals = {k: 0.0 for k in NUTRITION_KEYS}
        for entry in menu.items:
            if entry.item_type == "ingredient":
                ing = _find_ing(entry.item_name, ingredients)
                if ing:
                    total_g += entry.amount
                    for k, v in ing.scaled_nutrition(entry.amount).items():
                        totals[k] += v
            elif entry.item_type == "meal":
                m = meals.get(entry.item_name)
                if m:
                    mg, mt = NutritionCalc.meal_totals(m, ingredients)
                    ratio = entry.amount / mg if mg > 0 else 0
                    total_g += entry.amount
                    for k in NUTRITION_KEYS:
                        totals[k] += mt[k] * ratio
        return round(total_g, 1), {k: round(v, 2) for k, v in totals.items()}


def _find_ing(name, ingredients):
    nl = name.lower()
    for ing in ingredients.values():
        if ing.name.lower() == nl:
            return ing
    return None


# ━━━━━━━━━━━━━━━━━━ TABLE HELPERS ━━━━━━━━━━━━━━━━━━━━

def _calorie_text_color(kcal_per100: float) -> QColor:
    if kcal_per100 < 150: return CAL_TEXT_GREEN
    elif kcal_per100 < 400: return CAL_TEXT_YELLOW
    else: return CAL_TEXT_ORANGE


class NumericItem(QTableWidgetItem):
    def __lt__(self, other):
        v1 = self.data(Qt.UserRole)
        v2 = other.data(Qt.UserRole) if other else None
        if v1 is not None and v2 is not None:
            try: return float(v1) < float(v2)
            except (TypeError, ValueError): pass
        return super().__lt__(other)


def _num_item(value, suffix=""):
    if isinstance(value, float):
        text = f"{value:.2f}" if value != int(value) else str(int(value))
    else:
        text = str(value)
    item = NumericItem(text + suffix)
    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    try: item.setData(Qt.UserRole, float(value))
    except (ValueError, TypeError): pass
    return item


def _color_row_text(table, row, color):
    for col in range(table.columnCount()):
        item = table.item(row, col)
        if item: item.setForeground(color)


def _make_total_item(text, is_label=False):
    item = QTableWidgetItem(str(text))
    item.setFlags(Qt.ItemIsEnabled)
    item.setBackground(QColor(C_TOTAL_BG))
    item.setForeground(QColor("#7abaff"))
    f = item.font(); f.setBold(True); f.setPointSize(12); item.setFont(f)
    item.setTextAlignment((Qt.AlignLeft if is_label else Qt.AlignRight) | Qt.AlignVCenter)
    return item


def _make_per100_item(text, is_label=False):
    item = QTableWidgetItem(str(text))
    item.setFlags(Qt.ItemIsEnabled)
    item.setBackground(QColor(C_PER100_BG))
    item.setForeground(QColor(C_TEXT2))
    f = item.font(); f.setItalic(True); item.setFont(f)
    item.setTextAlignment((Qt.AlignLeft if is_label else Qt.AlignRight) | Qt.AlignVCenter)
    return item


def _fmt(v):
    if v == 0: return "0"
    if v == int(v): return str(int(v))
    return f"{v:.2f}"


# ━━━━━━━━━━━━━━━━━━━━━ DIALOGS ━━━━━━━━━━━━━━━━━━━━━━━

class IngredientDialog(QDialog):
    def __init__(self, parent=None, ingredient=None, next_id=1,
                 existing=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Ingredient" if ingredient else "New Ingredient")
        self.setMinimumWidth(480)
        self.ingredient = ingredient
        lay = QFormLayout(self)
        lay.setSpacing(12); lay.setContentsMargins(24, 24, 24, 24)
        names = sorted(set(i.name for i in (existing or {}).values()))
        brands = sorted(set(i.brand for i in (existing or {}).values() if i.brand))
        products = sorted(set(i.product_name for i in (existing or {}).values() if i.product_name))
        sizes = sorted(set(i.package_size for i in (existing or {}).values() if i.package_size))
        self.name_edit = QLineEdit(ingredient.name if ingredient else "")
        self.name_edit.setPlaceholderText("e.g. Chicken breast")
        self._add_completer(self.name_edit, names)
        self.brand_edit = QLineEdit(ingredient.brand if ingredient else "")
        self.brand_edit.setPlaceholderText("e.g. Organic Farm")
        self._add_completer(self.brand_edit, brands)
        self.product_edit = QLineEdit(ingredient.product_name if ingredient else "")
        self.product_edit.setPlaceholderText("e.g. Bio Hähnchenbrust")
        self._add_completer(self.product_edit, products)
        self.package_edit = QLineEdit(ingredient.package_size if ingredient else "")
        self.package_edit.setPlaceholderText("e.g. 500")
        self._add_completer(self.package_edit, sizes)
        lay.addRow("Name *", self.name_edit)
        lay.addRow("Brand", self.brand_edit)
        lay.addRow("Product Name", self.product_edit)
        lay.addRow("Package Size", self.package_edit)
        sep = QFrame(); sep.setObjectName("separator"); sep.setFrameShape(QFrame.HLine)
        lay.addRow(sep)
        lbl = QLabel("Nutrition per 100 g")
        lbl.setStyleSheet(f"font-weight: 700; font-size: 14px; color: {C_ACCENT};")
        lay.addRow(lbl)
        self.spins: Dict[str, QDoubleSpinBox] = {}
        for key, label in NUTRITION_FIELDS:
            s = QDoubleSpinBox()
            s.setRange(0, 99999); s.setDecimals(2); s.setSingleStep(0.1)
            if ingredient: s.setValue(getattr(ingredient, key))
            self.spins[key] = s
            lay.addRow(label, s)
        btns = QHBoxLayout(); btns.addStretch()
        cancel = QPushButton("Cancel"); cancel.setProperty("class", "secondary")
        cancel.clicked.connect(self.reject)
        save = QPushButton("Save"); save.clicked.connect(self.accept); save.setDefault(True)
        btns.addWidget(cancel); btns.addWidget(save)
        lay.addRow(btns)
        self._next_id = next_id

    @staticmethod
    def _add_completer(edit, items):
        c = QCompleter(items)
        c.setCaseSensitivity(Qt.CaseInsensitive)
        c.setFilterMode(Qt.MatchContains)
        c.popup().setStyleSheet(
            f"background: #3a3a3c; color: {C_TEXT}; border: 1px solid {C_BORDER2};"
            f"selection-background-color: {C_ACCENT}; outline: none; font-size: 13px;")
        edit.setCompleter(c)

    def get_ingredient(self):
        name = self.name_edit.text().strip()
        if not name: return None
        return Ingredient(
            id=self.ingredient.id if self.ingredient else self._next_id,
            name=name, brand=self.brand_edit.text().strip(),
            product_name=self.product_edit.text().strip(),
            package_size=self.package_edit.text().strip(),
            **{k: s.value() for k, s in self.spins.items()})


class IngredientPickerDialog(QDialog):
    """Unified picker for ingredients (and optionally meals).
    Used by BOTH Meals and Menus tabs — 100% identical appearance.
    Field order: Name → Brand (auto-filled dropdown) → Branded Product Name (auto-filled dropdown) → Amount (g).
    For meals: Brand empty, Branded Product Name = 'Homemade Food'."""

    def __init__(self, parent, ingredients, meals=None):
        super().__init__(parent)
        self.setWindowTitle("Select Ingredient")
        self.setMinimumWidth(520)
        self.ingredients = ingredients
        self._all_ings = sorted(ingredients.values(), key=lambda i: i.name.lower())
        self._meals = meals or {}
        self._meal_names = sorted(self._meals.keys())

        lay = QFormLayout(self)
        lay.setSpacing(12); lay.setContentsMargins(24, 24, 24, 24)

        # 1. NAME — prefixed with [Ingredient] or [Meal]
        self.name_combo = QComboBox()
        self.name_combo.setEditable(True)
        self.name_combo.setInsertPolicy(QComboBox.NoInsert)
        self.name_combo.currentTextChanged.connect(self._on_name_changed)
        lay.addRow("Name", self.name_combo)

        # 2. BRAND (dropdown, auto-filled when name is selected)
        self.brand_combo = QComboBox()
        self.brand_combo.setEditable(True)
        self.brand_combo.setInsertPolicy(QComboBox.NoInsert)
        lay.addRow("Brand", self.brand_combo)

        # 3. BRANDED PRODUCT NAME (dropdown, auto-filled)
        self.product_combo = QComboBox()
        self.product_combo.setEditable(True)
        self.product_combo.setInsertPolicy(QComboBox.NoInsert)
        lay.addRow("Branded Product Name", self.product_combo)

        # 4. AMOUNT (always grams)
        self.amount_spin = QDoubleSpinBox()
        self.amount_spin.setRange(0.1, 99999); self.amount_spin.setDecimals(1)
        self.amount_spin.setValue(100)
        lay.addRow("Weight (g)", self.amount_spin)

        btns = QHBoxLayout(); btns.addStretch()
        cancel = QPushButton("Cancel"); cancel.setProperty("class", "secondary"); cancel.clicked.connect(self.reject)
        add = QPushButton("Add"); add.clicked.connect(self.accept); add.setDefault(True)
        btns.addWidget(cancel); btns.addWidget(add)
        lay.addRow(btns)

        self._populate_names()

        # Start with empty name field, placeholder, and focus
        self._auto_skip = False  # set True when unique ingredient detected
        self.name_combo.setCurrentText("")
        self.name_combo.lineEdit().setPlaceholderText("e.g. Chicken breast")
        self.name_combo.lineEdit().installEventFilter(self)
        self.brand_combo.clear()
        self.product_combo.clear()
        QTimer.singleShot(0, lambda: self.name_combo.setFocus())

    def eventFilter(self, obj, event):
        """Skip to Amount only when user confirms name with Tab or Enter."""
        if obj == self.name_combo.lineEdit() and event.type() == QEvent.KeyPress:
            if event.key() in (Qt.Key_Tab, Qt.Key_Return, Qt.Key_Enter):
                # Accept completer selection first
                comp = self.name_combo.completer()
                if comp and comp.popup() and comp.popup().isVisible():
                    idx = comp.popup().currentIndex()
                    if idx.isValid():
                        self.name_combo.setCurrentText(idx.data())
                    comp.popup().hide()
                # Skip to Amount if auto-fill is complete
                if self._auto_skip:
                    QTimer.singleShot(50, self.amount_spin.setFocus)
                    QTimer.singleShot(60, self.amount_spin.selectAll)
                    if event.key() == Qt.Key_Tab:
                        return True
        return super().eventFilter(obj, event)

    @staticmethod
    def _set_completer(combo, items):
        c = QCompleter(items)
        c.setCaseSensitivity(Qt.CaseInsensitive); c.setFilterMode(Qt.MatchContains)
        c.popup().setStyleSheet(
            f"background: #3a3a3c; color: {C_TEXT}; border: 1px solid {C_BORDER2};"
            f"selection-background-color: {C_ACCENT}; outline: none; font-size: 13px;")
        combo.setCompleter(c)

    def _populate_names(self):
        self.name_combo.blockSignals(True)
        self.name_combo.clear()
        self._name_items = []
        for mn in self._meal_names:
            self._name_items.append((f"[Meal]  {mn}", "meal", mn))
        for ing in self._all_ings:
            self._name_items.append((f"[Ingredient]  {ing.name}", "ingredient", ing.name))
        display = [it[0] for it in self._name_items]
        self.name_combo.addItems(display)
        self._set_completer(self.name_combo, display)
        self.name_combo.blockSignals(False)

    def _on_name_changed(self, text):
        """When name changes, auto-fill Brand and Branded Product Name."""
        text = text.strip()
        self._auto_skip = False
        self.brand_combo.blockSignals(True)
        self.product_combo.blockSignals(True)
        self.brand_combo.clear()
        self.product_combo.clear()

        # Check if it's a meal
        for display, itype, name in self._name_items:
            if (display == text or name == text) and itype == "meal":
                self.brand_combo.addItem("Homemade Food")
                self.product_combo.addItem("Homemade Food")
                self.brand_combo.blockSignals(False)
                self.product_combo.blockSignals(False)
                self._auto_skip = True  # will skip on Tab/Enter
                return

        # Find all ingredients matching this name
        matching = []
        for display, itype, name in self._name_items:
            if (display == text or name.lower() == text.lower() or
                text.lower().endswith(name.lower())) and itype == "ingredient":
                matching = [i for i in self._all_ings if i.name == name]
                break
        if not matching:
            matching = [i for i in self._all_ings if i.name.lower() == text.lower()]

        if matching:
            brands = list(dict.fromkeys((i.brand or "") for i in matching))
            products = list(dict.fromkeys((i.product_name or "") for i in matching))
            brands = [b if b else "No brand" for b in brands]
            products = [p if p else "Whole food" for p in products]
            self.brand_combo.addItems(brands)
            self._set_completer(self.brand_combo, brands)
            self.product_combo.addItems(products)
            self._set_completer(self.product_combo, products)
            # Flag auto-skip if unique (single brand+product)
            if len(matching) == 1:
                self._auto_skip = True
        else:
            self.brand_combo.addItem("")
            self.product_combo.addItem("")

        self.brand_combo.blockSignals(False)
        self.product_combo.blockSignals(False)

    def get_result(self):
        text = self.name_combo.currentText().strip()
        amount = self.amount_spin.value()
        for display, itype, name in self._name_items:
            if display == text or name.lower() == text.lower():
                if itype == "meal":
                    return MenuEntry("meal", name, amount)
                else:
                    for ing in self._all_ings:
                        if ing.name == name:
                            return (ing.id, amount)
                    break
        for ing in self._all_ings:
            if ing.name.lower() == text.lower():
                return (ing.id, amount)
        return None



# ━━━━━━━━━━━━━━━━━ INGREDIENTS TAB ━━━━━━━━━━━━━━━━━━━

class IngredientsTab(QWidget):
    data_changed = Signal()

    def __init__(self):
        super().__init__()
        self.ingredients: Dict[int, Ingredient] = {}
        self.db_headers: List[str] = []
        self._col_map: Dict[str, int] = {}  # field → 0-based column index
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20); lay.setSpacing(14)
        hdr = QHBoxLayout()
        t = QLabel("Ingredients"); t.setObjectName("sectionTitle")
        hdr.addWidget(t); hdr.addStretch()
        self.search = QLineEdit()
        self.search.setPlaceholderText("Search…"); self.search.setFixedWidth(280)
        self.search.textChanged.connect(self._filter)
        hdr.addWidget(self.search)
        add = QPushButton("＋  Add Ingredient"); add.clicked.connect(self._add)
        hdr.addWidget(add)
        lay.addLayout(hdr)
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        # ── Calorie color legend ──
        legend = QHBoxLayout()
        legend.addStretch()
        for color, label in [
            (CAL_TEXT_GREEN,  "● 0–149 kcal (Low)"),
            (CAL_TEXT_YELLOW, "● 150–399 kcal (Medium)"),
            (CAL_TEXT_ORANGE, "● ≥ 400 kcal (High)"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(
                f"color: rgb({color.red()},{color.green()},{color.blue()});"
                f"font-size: 11px; font-weight: 600; padding: 2px 12px; background: transparent;"
            )
            legend.addWidget(lbl)
        legend.addStretch()
        lay.addLayout(legend)

        btns = QHBoxLayout(); btns.addStretch()
        eb = QPushButton("Edit"); eb.setProperty("class", "secondary"); eb.clicked.connect(self._edit)
        db = QPushButton("Delete"); db.setProperty("class", "danger"); db.clicked.connect(self._delete)
        btns.addWidget(eb); btns.addWidget(db)
        lay.addLayout(btns)

    def set_headers(self, headers: List[str], col_map: Dict[str, int] = None):
        """Set column headers exactly as read from the database."""
        self.db_headers = headers
        if col_map:
            self._col_map = col_map
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

    def load(self, ingredients):
        self.ingredients = ingredients; self._populate()

    def _populate(self, ft=""):
        ft = ft.lower()
        self.table.setSortingEnabled(False)
        items = [i for i in self.ingredients.values()
                 if not ft or ft in i.name.lower() or ft in i.brand.lower()]
        n_cols = len(self.db_headers) if self.db_headers else 16

        # Use detected col_map, or fallback to standard positions
        cm = self._col_map if self._col_map else {
            "energy_kj": 6, "energy_kcal": 7, "fat": 8,
            "saturated_fat": 9, "carbohydrate": 10, "sugars": 11,
            "fibre": 12, "protein": 13, "salt": 14,
        }

        self.table.setRowCount(len(items))
        for r, ing in enumerate(sorted(items, key=lambda x: x.name.lower())):
            # Fill all columns with empty items first
            for c in range(n_cols):
                self.table.setItem(r, c, QTableWidgetItem(""))
            # Fixed columns
            self.table.setItem(r, 0, _num_item(ing.id))
            self.table.setItem(r, 2, QTableWidgetItem(ing.name))
            self.table.setItem(r, 3, QTableWidgetItem(ing.brand))
            self.table.setItem(r, 4, QTableWidgetItem(ing.product_name))
            self.table.setItem(r, 5, QTableWidgetItem("per 100 g"))
            if n_cols > 15:
                self.table.setItem(r, n_cols - 1, QTableWidgetItem(ing.package_size))
            # Nutrition at detected column positions
            for field in NUTRITION_KEYS:
                col = cm.get(field, -1)
                if 0 <= col < n_cols:
                    self.table.setItem(r, col, _num_item(getattr(ing, field)))
            # Calorie color
            _color_row_text(self.table, r, _calorie_text_color(ing.energy_kcal))
        self.table.resizeColumnsToContents()
        self.table.setSortingEnabled(True)

    def _filter(self, text): self._populate(text)
    def _next_id(self): return max(self.ingredients.keys(), default=0) + 1

    def _add(self):
        d = IngredientDialog(self, next_id=self._next_id(), existing=self.ingredients)
        if d.exec() == QDialog.Accepted:
            ing = d.get_ingredient()
            if ing:
                self.ingredients[ing.id] = ing
                self._populate(self.search.text()); self.data_changed.emit()

    def _edit(self):
        row = self.table.currentRow()
        if row < 0: return
        iid = int(self.table.item(row, 0).text())
        ing = self.ingredients.get(iid)
        if not ing: return
        d = IngredientDialog(self, ingredient=ing, existing=self.ingredients)
        if d.exec() == QDialog.Accepted:
            u = d.get_ingredient()
            if u:
                self.ingredients[u.id] = u
                self._populate(self.search.text()); self.data_changed.emit()

    def _delete(self):
        row = self.table.currentRow()
        if row < 0: return
        iid = int(self.table.item(row, 0).text())
        name = self.table.item(row, 2).text()  # col C = name
        if QMessageBox.question(self, "Delete Ingredient",
            f"Delete '{name}'?\nIt will be removed from all meals using it.") == QMessageBox.Yes:
            del self.ingredients[iid]
            self._populate(self.search.text()); self.data_changed.emit()


# ━━━━━━━━━━━━━━━━━━━ MEALS TAB ━━━━━━━━━━━━━━━━━━━━━━

LEFT_PANEL_WIDTH = 320

class MealsTab(QWidget):
    data_changed = Signal()

    def __init__(self, ing_tab):
        super().__init__()
        self.ing_tab = ing_tab; self.meals: Dict[str, Meal] = {}
        self._nutr_keys = NUTRITION_KEYS      # updated from DB
        self._nutr_labels = NUTRITION_LABELS
        self._build()

    @property
    def ingredients(self): return self.ing_tab.ingredients

    def set_nutrition_order(self, keys, labels):
        self._nutr_keys = keys
        self._nutr_labels = labels
        cols = ["Ingredient", "Weight (g)"] + labels
        self.detail.setColumnCount(len(cols))
        self.detail.setHorizontalHeaderLabels(cols)

    def _build(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20); lay.setSpacing(14)
        left = QVBoxLayout(); left.setSpacing(10)
        t = QLabel("Meals"); t.setObjectName("sectionTitle"); left.addWidget(t)
        self.search = QLineEdit(); self.search.setPlaceholderText("Search…")
        self.search.textChanged.connect(self._filter_list); left.addWidget(self.search)
        self.meal_list = QListWidget()
        self.meal_list.currentItemChanged.connect(self._on_select)
        left.addWidget(self.meal_list)
        lb = QHBoxLayout(); lb.setSpacing(6)
        a = QPushButton("＋ New"); a.clicked.connect(self._add_meal)
        r = QPushButton("Rename"); r.setProperty("class", "secondary"); r.clicked.connect(self._rename_meal)
        d = QPushButton("Delete"); d.setProperty("class", "danger"); d.clicked.connect(self._delete_meal)
        lb.addWidget(a); lb.addWidget(r); lb.addWidget(d)
        left.addLayout(lb)
        lw = QWidget(); lw.setLayout(left); lw.setFixedWidth(LEFT_PANEL_WIDTH)
        lay.addWidget(lw)
        right = QVBoxLayout(); right.setSpacing(10)
        self.title_lbl = QLabel("Select a meal"); self.title_lbl.setObjectName("sectionTitle")
        right.addWidget(self.title_lbl)
        self.detail = QTableWidget()
        self.detail.setAlternatingRowColors(True)
        self.detail.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.detail.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.detail.verticalHeader().setVisible(False)
        self.detail.horizontalHeader().setStretchLastSection(True)
        cols = ["Ingredient", "Weight (g)"] + NUTRITION_LABELS
        self.detail.setColumnCount(len(cols))
        self.detail.setHorizontalHeaderLabels(cols)
        right.addWidget(self.detail)
        rb = QHBoxLayout()
        ai = QPushButton("＋ Add Ingredient"); ai.clicked.connect(self._add_ing)
        ea = QPushButton("Edit Weight"); ea.setProperty("class", "secondary"); ea.clicked.connect(self._edit_amount)
        ri = QPushButton("Remove"); ri.setProperty("class", "danger"); ri.clicked.connect(self._remove_ing)
        rb.addWidget(ai); rb.addWidget(ea); rb.addStretch(); rb.addWidget(ri)
        right.addLayout(rb)
        lay.addLayout(right, 1)

    def load(self, meals): self.meals = meals; self._populate_list()
    def _populate_list(self, ft=""):
        ft = ft.lower()
        self.meal_list.blockSignals(True)
        cur = self.meal_list.currentItem()
        cn = cur.text() if cur else None
        self.meal_list.clear()
        for n in sorted(self.meals, key=str.lower):
            if ft and ft not in n.lower(): continue
            self.meal_list.addItem(n)
        if cn:
            found = self.meal_list.findItems(cn, Qt.MatchExactly)
            if found: self.meal_list.setCurrentItem(found[0])
        self.meal_list.blockSignals(False)
        if self.meal_list.currentItem(): self._on_select(self.meal_list.currentItem())

    def _filter_list(self, t): self._populate_list(t)
    def _cur(self):
        i = self.meal_list.currentItem()
        return self.meals.get(i.text()) if i else None

    def _on_select(self, cur, prev=None):
        meal = self._cur()
        if not meal: self.title_lbl.setText("Select a meal"); self.detail.setRowCount(0); return
        self.title_lbl.setText(meal.name); self._refresh()

    def _refresh(self):
        meal = self._cur()
        if not meal: return
        n = len(meal.items)
        nk = self._nutr_keys
        self.detail.setSortingEnabled(False)
        self.detail.setRowCount(n + 2)
        for r, mi in enumerate(meal.items):
            ing = self.ingredients.get(mi.ingredient_id)
            self.detail.setItem(r, 0, QTableWidgetItem(ing.name if ing else f"[ID {mi.ingredient_id}]"))
            self.detail.setItem(r, 1, _num_item(mi.amount_grams))
            if ing:
                sc = ing.scaled_nutrition(mi.amount_grams)
                for c, key in enumerate(nk):
                    self.detail.setItem(r, 2 + c, _num_item(sc[key]))
                _color_row_text(self.detail, r, _calorie_text_color(ing.energy_kcal))
        tg, tt = NutritionCalc.meal_totals(meal, self.ingredients)
        self.detail.setItem(n, 0, _make_total_item("TOTAL", True))
        self.detail.setItem(n, 1, _make_total_item(_fmt(tg)))
        for c, key in enumerate(nk):
            self.detail.setItem(n, 2 + c, _make_total_item(_fmt(tt[key])))
        p = NutritionCalc.per100(tg, tt)
        self.detail.setItem(n+1, 0, _make_per100_item("per 100 g", True))
        self.detail.setItem(n+1, 1, _make_per100_item("100"))
        for c, key in enumerate(nk):
            self.detail.setItem(n+1, 2 + c, _make_per100_item(_fmt(p[key])))
        self.detail.resizeColumnsToContents()

    def _add_meal(self):
        n, ok = QInputDialog.getText(self, "New Meal", "Meal name:")
        n = n.strip() if ok else ""
        if not n: return
        if n in self.meals: QMessageBox.warning(self, "Exists", f"'{n}' already exists."); return
        self.meals[n] = Meal(name=n); self._populate_list()
        found = self.meal_list.findItems(n, Qt.MatchExactly)
        if found: self.meal_list.setCurrentItem(found[0])
        self.data_changed.emit()

    def _rename_meal(self):
        m = self._cur()
        if not m: return
        nn, ok = QInputDialog.getText(self, "Rename", "New name:", text=m.name)
        nn = nn.strip() if ok else ""
        if not nn or nn == m.name: return
        if nn in self.meals: QMessageBox.warning(self, "Exists", f"'{nn}' already exists."); return
        del self.meals[m.name]; m.name = nn; self.meals[nn] = m; self._populate_list()
        found = self.meal_list.findItems(nn, Qt.MatchExactly)
        if found: self.meal_list.setCurrentItem(found[0])
        self.data_changed.emit()

    def _delete_meal(self):
        m = self._cur()
        if not m: return
        if QMessageBox.question(self, "Delete", f"Delete meal '{m.name}'?") == QMessageBox.Yes:
            del self.meals[m.name]; self._populate_list(); self.data_changed.emit()

    def _add_ing(self):
        m = self._cur()
        if not m: QMessageBox.information(self, "No Meal", "Select or create a meal first."); return
        d = IngredientPickerDialog(self, self.ingredients)
        if d.exec() == QDialog.Accepted:
            r = d.get_result()
            if r and isinstance(r, tuple):
                m.items.append(MealIngredient(ingredient_id=r[0], amount_grams=r[1]))
                self._refresh(); self.data_changed.emit()

    def _edit_amount(self):
        m = self._cur(); row = self.detail.currentRow()
        if not m or row < 0 or row >= len(m.items): return
        mi = m.items[row]
        amt, ok = QInputDialog.getDouble(self, "Edit Weight", "Weight (g):", mi.amount_grams, 0.1, 99999, 1)
        if ok: mi.amount_grams = amt; self._refresh(); self.data_changed.emit()

    def _remove_ing(self):
        m = self._cur(); row = self.detail.currentRow()
        if not m or row < 0 or row >= len(m.items): return
        m.items.pop(row); self._refresh(); self.data_changed.emit()


# ━━━━━━━━━━━━━━━━━━━ MENUS TAB ━━━━━━━━━━━━━━━━━━━━━━

class MenusTab(QWidget):
    data_changed = Signal()

    def __init__(self, ing_tab, meals_tab):
        super().__init__()
        self.ing_tab = ing_tab; self.meals_tab = meals_tab
        self.menus: Dict[str, Menu] = {}
        self._nutr_keys = NUTRITION_KEYS
        self._nutr_labels = NUTRITION_LABELS
        self._build()

    @property
    def ingredients(self): return self.ing_tab.ingredients
    @property
    def meals(self): return self.meals_tab.meals

    def set_nutrition_order(self, keys, labels):
        self._nutr_keys = keys
        self._nutr_labels = labels
        cols = ["Type", "Name", "Weight (g)"] + labels
        self.detail.setColumnCount(len(cols))
        self.detail.setHorizontalHeaderLabels(cols)

    def _build(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20); lay.setSpacing(14)
        left = QVBoxLayout(); left.setSpacing(10)
        t = QLabel("Menus"); t.setObjectName("sectionTitle"); left.addWidget(t)
        self.search = QLineEdit(); self.search.setPlaceholderText("Search…")
        self.search.textChanged.connect(self._filter_list); left.addWidget(self.search)
        self.menu_list = QListWidget()
        self.menu_list.currentItemChanged.connect(self._on_select)
        left.addWidget(self.menu_list)
        lb = QHBoxLayout(); lb.setSpacing(6)
        a = QPushButton("＋ New"); a.clicked.connect(self._add_menu)
        r = QPushButton("Rename"); r.setProperty("class", "secondary"); r.clicked.connect(self._rename_menu)
        d = QPushButton("Delete"); d.setProperty("class", "danger"); d.clicked.connect(self._delete_menu)
        lb.addWidget(a); lb.addWidget(r); lb.addWidget(d)
        left.addLayout(lb)
        lw = QWidget(); lw.setLayout(left); lw.setFixedWidth(LEFT_PANEL_WIDTH)
        lay.addWidget(lw)
        right = QVBoxLayout(); right.setSpacing(10)
        self.title_lbl = QLabel("Select a menu"); self.title_lbl.setObjectName("sectionTitle")
        right.addWidget(self.title_lbl)
        self.detail = QTableWidget()
        self.detail.setAlternatingRowColors(True)
        self.detail.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.detail.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.detail.verticalHeader().setVisible(False)
        self.detail.horizontalHeader().setStretchLastSection(True)
        cols = ["Type", "Name", "Weight (g)"] + NUTRITION_LABELS
        self.detail.setColumnCount(len(cols)); self.detail.setHorizontalHeaderLabels(cols)
        right.addWidget(self.detail)
        rb = QHBoxLayout()
        ai = QPushButton("＋ Add Item"); ai.clicked.connect(self._add_item)
        ea = QPushButton("Edit Weight"); ea.setProperty("class", "secondary"); ea.clicked.connect(self._edit_amount)
        ri = QPushButton("Remove"); ri.setProperty("class", "danger"); ri.clicked.connect(self._remove_item)
        rb.addWidget(ai); rb.addWidget(ea); rb.addStretch(); rb.addWidget(ri)
        right.addLayout(rb)
        lay.addLayout(right, 1)

    def load(self, menus): self.menus = menus; self._populate_list()
    def _populate_list(self, ft=""):
        ft = ft.lower()
        self.menu_list.blockSignals(True); cur = self.menu_list.currentItem()
        cn = cur.text() if cur else None; self.menu_list.clear()
        for n in sorted(self.menus, key=str.lower):
            if ft and ft not in n.lower(): continue
            self.menu_list.addItem(n)
        if cn:
            found = self.menu_list.findItems(cn, Qt.MatchExactly)
            if found: self.menu_list.setCurrentItem(found[0])
        self.menu_list.blockSignals(False)
        if self.menu_list.currentItem(): self._on_select(self.menu_list.currentItem())

    def _filter_list(self, t): self._populate_list(t)
    def _cur(self):
        i = self.menu_list.currentItem()
        return self.menus.get(i.text()) if i else None

    def _on_select(self, cur, prev=None):
        menu = self._cur()
        if not menu: self.title_lbl.setText("Select a menu"); self.detail.setRowCount(0); return
        self.title_lbl.setText(menu.name); self._refresh()

    def _refresh(self):
        menu = self._cur()
        if not menu: return
        ni = len(menu.items)
        nk = self._nutr_keys
        self.detail.setSortingEnabled(False); self.detail.setRowCount(ni + 2)
        for r, entry in enumerate(menu.items):
            self.detail.setItem(r, 0, QTableWidgetItem(entry.item_type.capitalize()))
            self.detail.setItem(r, 1, QTableWidgetItem(entry.item_name))
            self.detail.setItem(r, 2, _num_item(entry.amount))
            kcal_for_color = 0.0
            if entry.item_type == "ingredient":
                ing = _find_ing(entry.item_name, self.ingredients)
                if ing:
                    kcal_for_color = ing.energy_kcal
                    sc = ing.scaled_nutrition(entry.amount)
                    for c, key in enumerate(nk):
                        self.detail.setItem(r, 3 + c, _num_item(sc[key]))
            else:
                meal = self.meals.get(entry.item_name)
                if meal:
                    mg, mt = NutritionCalc.meal_totals(meal, self.ingredients)
                    ratio = entry.amount / mg if mg > 0 else 0
                    for c, key in enumerate(nk):
                        self.detail.setItem(r, 3 + c, _num_item(round(mt[key] * ratio, 2)))
                    p100 = NutritionCalc.per100(mg, mt)
                    kcal_for_color = p100.get("energy_kcal", 0)
            _color_row_text(self.detail, r, _calorie_text_color(kcal_for_color))
        tg, tt = NutritionCalc.menu_totals(menu, self.meals, self.ingredients)
        self.detail.setItem(ni, 0, _make_total_item("TOTAL", True))
        self.detail.setItem(ni, 1, _make_total_item(""))
        self.detail.setItem(ni, 2, _make_total_item(_fmt(tg)))
        for c, key in enumerate(nk):
            self.detail.setItem(ni, 3 + c, _make_total_item(_fmt(tt[key])))
        p = NutritionCalc.per100(tg, tt)
        self.detail.setItem(ni+1, 0, _make_per100_item("per 100 g", True))
        self.detail.setItem(ni+1, 1, _make_per100_item(""))
        self.detail.setItem(ni+1, 2, _make_per100_item("100"))
        for c, key in enumerate(nk):
            self.detail.setItem(ni+1, 3 + c, _make_per100_item(_fmt(p[key])))
        self.detail.resizeColumnsToContents()

    def _add_menu(self):
        n, ok = QInputDialog.getText(self, "New Menu", "Menu name:")
        n = n.strip() if ok else ""
        if not n: return
        if n in self.menus: QMessageBox.warning(self, "Exists", f"'{n}' already exists."); return
        self.menus[n] = Menu(name=n); self._populate_list()
        found = self.menu_list.findItems(n, Qt.MatchExactly)
        if found: self.menu_list.setCurrentItem(found[0])
        self.data_changed.emit()

    def _rename_menu(self):
        m = self._cur()
        if not m: return
        nn, ok = QInputDialog.getText(self, "Rename", "New name:", text=m.name)
        nn = nn.strip() if ok else ""
        if not nn or nn == m.name: return
        if nn in self.menus: QMessageBox.warning(self, "Exists", f"'{nn}' already exists."); return
        del self.menus[m.name]; m.name = nn; self.menus[nn] = m; self._populate_list()
        found = self.menu_list.findItems(nn, Qt.MatchExactly)
        if found: self.menu_list.setCurrentItem(found[0])
        self.data_changed.emit()

    def _delete_menu(self):
        m = self._cur()
        if not m: return
        if QMessageBox.question(self, "Delete", f"Delete menu '{m.name}'?") == QMessageBox.Yes:
            del self.menus[m.name]; self._populate_list(); self.data_changed.emit()

    def _add_item(self):
        m = self._cur()
        if not m: QMessageBox.information(self, "No Menu", "Select or create a menu first."); return
        d = IngredientPickerDialog(self, self.ingredients, meals=self.meals)
        if d.exec() == QDialog.Accepted:
            result = d.get_result()
            if result is None:
                return
            if isinstance(result, MenuEntry):
                # Meal — amount is already in grams
                m.items.append(result)
            else:
                # Ingredient: result = (ing_id, amount_grams)
                ing_id, amount = result
                ing = self.ingredients.get(ing_id)
                if ing:
                    m.items.append(MenuEntry("ingredient", ing.name, amount))
            self._refresh(); self.data_changed.emit()

    def _edit_amount(self):
        m = self._cur(); row = self.detail.currentRow()
        if not m or row < 0 or row >= len(m.items): return
        entry = m.items[row]
        amt, ok = QInputDialog.getDouble(self, "Edit Weight", "Weight (g):", entry.amount, 0.01, 99999, 2)
        if ok: entry.amount = amt; self._refresh(); self.data_changed.emit()

    def _remove_item(self):
        m = self._cur(); row = self.detail.currentRow()
        if not m or row < 0 or row >= len(m.items): return
        m.items.pop(row); self._refresh(); self.data_changed.emit()


# ━━━━━━━━━━━━━━━━━━ MAIN WINDOW ━━━━━━━━━━━━━━━━━━━━━

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE); self.resize(1240, 780)
        self.db = None
        self.settings = QSettings(SETTINGS_ORG, SETTINGS_APP)
        central = QWidget(); self.setCentralWidget(central)
        self.root = QVBoxLayout(central)
        self.root.setContentsMargins(0, 0, 0, 0); self.root.setSpacing(0)
        self.db_bar = QFrame(); self.db_bar.setObjectName("dbBar"); self.db_bar.setFixedHeight(52)
        db_lay = QHBoxLayout(self.db_bar)
        db_lay.setContentsMargins(20, 0, 20, 0); db_lay.setSpacing(12)
        self.db_icon = QLabel("◉")
        self.db_icon.setStyleSheet(f"font-size: 16px; color: {C_TEXT3}; background: transparent;")
        db_lay.addWidget(self.db_icon)
        db_label = QLabel("Database")
        db_label.setStyleSheet(f"color: {C_TEXT2}; font-size: 12px; font-weight: 600; background: transparent;")
        db_lay.addWidget(db_label)
        self.db_path_label = QLabel("No database selected"); self.db_path_label.setObjectName("dbPath")
        db_lay.addWidget(self.db_path_label, 1)
        browse_btn = QPushButton("Change…")
        browse_btn.setProperty("class", "dbButton"); browse_btn.setFixedWidth(100)
        browse_btn.clicked.connect(self._browse_db)
        db_lay.addWidget(browse_btn)
        self.root.addWidget(self.db_bar)
        self.ing_tab = IngredientsTab()
        self.meals_tab = MealsTab(self.ing_tab)
        self.menus_tab = MenusTab(self.ing_tab, self.meals_tab)
        self.tabs = QTabWidget()
        self.tabs.addTab(self.ing_tab, "   Ingredients   ")
        self.tabs.addTab(self.meals_tab, "   Meals   ")
        self.tabs.addTab(self.menus_tab, "   Menus   ")
        self.tabs.currentChanged.connect(self._on_tab)
        self.root.addWidget(self.tabs, 1)
        self.status = QStatusBar(); self.setStatusBar(self.status)
        self._save_timer = QTimer()
        self._save_timer.setSingleShot(True); self._save_timer.setInterval(2000)
        self._save_timer.timeout.connect(self._do_save)
        self.ing_tab.data_changed.connect(self._schedule_save)
        self.meals_tab.data_changed.connect(self._schedule_save)
        self.menus_tab.data_changed.connect(self._schedule_save)
        mb = self.menuBar()
        mb.setStyleSheet(f"QMenuBar {{ background: {C_CARD}; color: {C_TEXT}; }}"
                         f"QMenuBar::item:selected {{ background: {C_BORDER2}; }}"
                         f"QMenu {{ background: #3a3a3c; color: {C_TEXT}; border: 1px solid {C_BORDER2}; }}"
                         f"QMenu::item:selected {{ background: {C_ACCENT}; }}")
        fm = mb.addMenu("File")
        sa = QAction("Save Now", self); sa.setShortcut("Ctrl+S"); sa.triggered.connect(self._do_save); fm.addAction(sa)
        fm.addSeparator()
        qa = QAction("Quit", self); qa.setShortcut("Ctrl+Q"); qa.triggered.connect(self.close); fm.addAction(qa)
        from PySide6.QtGui import QKeySequence, QShortcut
        add_sc = QShortcut(QKeySequence("Ctrl++"), self); add_sc.activated.connect(self._on_add_shortcut)
        add_sc2 = QShortcut(QKeySequence("Ctrl+="), self); add_sc2.activated.connect(self._on_add_shortcut)
        tab_fwd = QShortcut(QKeySequence("Ctrl+Tab"), self)
        tab_fwd.activated.connect(lambda: self.tabs.setCurrentIndex((self.tabs.currentIndex() + 1) % self.tabs.count()))
        tab_bwd = QShortcut(QKeySequence("Ctrl+Shift+Tab"), self)
        tab_bwd.activated.connect(lambda: self.tabs.setCurrentIndex((self.tabs.currentIndex() - 1) % self.tabs.count()))
        last = self.settings.value(SETTINGS_LAST_DB, "")
        if last and Path(last).is_file(): self._open_db(last)
        else: self.status.showMessage("Select a database to get started")

    def _browse_db(self):
        start = str(Path(self.db.path).parent) if self.db else str(Path.home())
        path, _ = QFileDialog.getOpenFileName(self, "Select Nutrition Database", start, "Excel Files (*.xlsx);;All Files (*)")
        if path: self._open_db(path)

    def _open_db(self, path):
        try:
            self.db = DatabaseManager(path)
            self.settings.setValue(SETTINGS_LAST_DB, path)

            # Load ingredients FIRST — this detects column positions from headers
            ingredients = self.db.load_ingredients()

            # NOW _col_map is populated — pass headers + col_map to UI
            all_headers = self.db.load_headers()
            ing_headers = all_headers.get(DatabaseManager.INGREDIENTS_SHEET, [])
            if ing_headers:
                self.ing_tab.set_headers(ing_headers, self.db._col_map)

            # Pass dynamic nutrition column order to detail tabs
            nutr_keys = self.db.nutrition_display_order
            nutr_labels = self.db.nutrition_display_labels
            self.meals_tab.set_nutrition_order(nutr_keys, nutr_labels)
            self.menus_tab.set_nutrition_order(nutr_keys, nutr_labels)

            self.ing_tab.load(ingredients)
            loaded_meals = self.db.load_meals()
            self.meals_tab.load(loaded_meals)
            self.menus_tab.load(self.db.load_menus(ingredients, loaded_meals))
            display = path if len(path) <= 65 else str(Path(Path(path).parts[0], "…", *Path(path).parts[-2:]))
            self.db_path_label.setText(display); self.db_path_label.setToolTip(path)
            self.db_icon.setStyleSheet(f"font-size: 16px; color: {C_GREEN}; background: transparent;")
            self.db_icon.setText("◉")
            self.status.showMessage(f"Loaded — {len(self.ing_tab.ingredients)} ingredients")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open database:\n\n{e}")

    def _schedule_save(self):
        if not self.db: return
        self.status.showMessage("Unsaved changes…"); self._save_timer.start()

    def _do_save(self):
        if not self.db: return
        try:
            self.db.save_all(self.ing_tab.ingredients, self.meals_tab.meals, self.menus_tab.menus)
            self.status.showMessage(f"Saved  ·  {datetime.now().strftime('%H:%M:%S')}")
        except Exception as e:
            self.status.showMessage(f"SAVE ERROR: {e}")
            QMessageBox.critical(self, "Save Error", f"Failed to save:\n\n{e}\n\nYour backup is intact.")

    def _on_tab(self, idx):
        if idx == 1: self.meals_tab._refresh()
        elif idx == 2: self.menus_tab._refresh()

    def _on_add_shortcut(self):
        idx = self.tabs.currentIndex()
        if idx == 0: self.ing_tab._add()
        elif idx == 1: self.meals_tab._add_ing()
        elif idx == 2: self.menus_tab._add_item()

    def closeEvent(self, event):
        if self.db:
            try:
                self.db.save_all(self.ing_tab.ingredients, self.meals_tab.meals, self.menus_tab.menus)
            except Exception as e:
                if QMessageBox.warning(self, "Save Error", f"Could not save:\n{e}\n\nQuit anyway?",
                                       QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
                    event.ignore(); return
        event.accept()


# ━━━━━━━━━━━━━━━━━━━━ ENTRY POINT ━━━━━━━━━━━━━━━━━━━━

def main():
    app = QApplication(sys.argv)
    app.setFont(QFontGui("SF Pro Text", 12))
    arrow_path = _create_arrow_image()
    app.setStyleSheet(_build_stylesheet(arrow_path))
    app.setApplicationName(SETTINGS_APP)
    app.setOrganizationName(SETTINGS_ORG)
    # Set application icon
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "..", "assets", "01_media", "01_icons", "icon.ico")
    if os.path.exists(icon_path):
        app_icon = QIcon(icon_path)
        app.setWindowIcon(app_icon)

    focus_filter = SelectAllOnFocus(app)
    app.installEventFilter(focus_filter)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()